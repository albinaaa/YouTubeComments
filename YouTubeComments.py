from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from bs4 import BeautifulSoup
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
import re


def ScrapComment(url):
    options = webdriver.ChromeOptions()
    options.add_argument('--ignore-certificate-errors-spki-list')
    options.add_argument("--headless")
    options.add_argument('--lang=en')
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36")
    s=Service('C:/Users/Albina/Downloads/chromedriver_win32/chromedriver.exe')
    driver = webdriver.Chrome(service=s, options=options)
    driver.get(url)
    prev_h = 0
    while True:
        height = driver.execute_script("""
                function getActualHeight() {
                    return Math.max(
                        Math.max(document.body.scrollHeight, document.documentElement.scrollHeight),
                        Math.max(document.body.offsetHeight, document.documentElement.offsetHeight),
                        Math.max(document.body.clientHeight, document.documentElement.clientHeight)
                    );
                }
                return getActualHeight();
            """)
        driver.execute_script(f"window.scrollTo({prev_h},{prev_h + 200})")
        # fix the time sleep value according to your network connection
        time.sleep(1)
        prev_h +=200  
        if prev_h >= height:
            break
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    driver.quit()
    title_text_div = soup.select_one('#container h1')
    title = title_text_div and title_text_div.text
    
    comment_auth = soup.select("#header-author #author-text > span")
    auth_list = [y.text for y in comment_auth]
    comment_div = soup.select("#content #content-text")
    comment_list = [x.text for x in comment_div]
    comment_url = soup.find_all("yt-formatted-string", class_="published-time-text style-scope ytd-comment-renderer")
    #find likes&comments
    clikes_parent = comment_url
    ccomments_parent = comment_url
    #end find likes&comments
    comment_urls=[]
    for c in comment_url:
        comment_urls.append(c.find("a", href=True))
    url_list = [y['href'] for y in comment_urls]

    comment_likes=[]
    for l in clikes_parent:
        #find action buttons
        ab = l.parent.parent.find_next(id="action-buttons")
        heart_action = ab.find_next(id="creator-heart")
        if heart_action.find(id="creator-heart-button") is not None:
            likes=1
        else:
            likes=0
        comment_likes.append(likes)
    
    replies = []
    for r in ccomments_parent:
        rp_main = r.parent.parent.parent
        rp = rp_main.parent.parent.find_next(id="replies")
        
        repl=-1
        if rp.get('hidden') is not None:
            repl=0
        else:
            rep=rp.find(class_="expander-header style-scope ytd-comment-replies-renderer")
            if rep.get('teaser') is not None:
                rrep=rep['teaser']
                if 'Олег Брагинский' in rrep:
                    repl=1
            else:
                repl=0
        replies.append(repl)


    all_comment = []
    for a in zip(comment_list, auth_list, url_list, comment_likes, replies):
        all_comment.append(list(a))
    
    return all_comment


if __name__ == "__main__":

    try:
        doc = openpyxl.reader.excel.load_workbook(filename="youtube.xlsx")
        print("loaded")

        new_sheet = doc.create_sheet('Comments')

        i=1
        new_sheet.cell(row=i, column=1).value = "Video ID"
        new_sheet.cell(row=i, column=2).value = "Text"
        new_sheet.cell(row=i, column=3).value = "Author"
        new_sheet.cell(row=i, column=4).value = "Comment URL"
        new_sheet.cell(row=i, column=5).value = "Like"
        new_sheet.cell(row=i, column=6).value = "Reply"
        new_sheet.cell(row=i, column=7).value = "Words count"
        new_sheet.cell(row=i, column=8).value = "Thanks"
        new_sheet.cell(row=i, column=9).value = "?"
        new_sheet.cell(row=i, column=10).value = ":"
        new_sheet.cell(row=i, column=11).value = ","
        new_sheet.cell(row=i, column=12).value = "!"



        doc.active = 0
        sheet = doc.worksheets[0]

        for row in sheet.iter_rows(min_row=2, min_col=1, max_row=2501, max_col=sheet.max_column):
            current_row=[]
            for cell in row:
                current_row.append(cell.value)
            num = current_row[0]
            url = current_row[1]    
            video_comments = ScrapComment(url)
            for c in video_comments:
                i = i+1
                n = new_sheet.cell(row=i, column=1)
                n.value = num
                
                c_comment = c[0].strip()

                com = new_sheet.cell(row=i, column=2)
                com.value = c_comment

                auth = new_sheet.cell(row=i, column=3)
                auth.value = c[1].strip()

                c_url = new_sheet.cell(row=i, column=4)
                c_url.value = "https://www.youtube.com"+c[2]

                c_like = new_sheet.cell(row=i, column=5)
                c_like.value = c[3]

                c_rep = new_sheet.cell(row=i, column=6)
                c_rep.value = c[4]

                w_count = new_sheet.cell(row=i, column=7)
                w_count.value = len(c_comment.split())

                th_count = new_sheet.cell(row=i, column=8)
                th_count.value = c_comment.lower().count('спасибо')

                q_count = new_sheet.cell(row=i, column=9)
                q_count.value = c[0].count('?')

                с_count = new_sheet.cell(row=i, column=10)
                с_count.value = c[0].count(':')

                comma_count = new_sheet.cell(row=i, column=11)
                comma_count.value = c[0].count(',')

                e_count = new_sheet.cell(row=i, column=12)
                e_count.value = c[0].count('!')

                print(num)
        doc.save('youtube.xlsx')
    except:
        doc.save('youtube.xlsx')
